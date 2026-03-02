#!/usr/bin/env python3
"""
convert_steam_data.py — One-time conversion of merged Excel to JSON

Input:  WSU_Energy_Data_FY21-FY26_Combined.xlsx
Output: data/steam_plant.json       (daily + monthly rollups)
        data/steam_plant_hourly.json (hourly data)

Run once. After that, the dashboard's import tool handles new months.
"""

import json
import os
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Column mappings ──────────────────────────────────────────────────────
# Maps Excel column index (0-based) to short key name
# Daily Data sheet: 71 columns (indices 0-70)

DAILY_COLS = {
    # Col 0: Day number (skip)
    # Col 1: Date
    # Boiler 1 (cols 2-6)
    2:  'b1_gas',           # hourly total gas (scf)
    3:  'b1_steam_gas',     # hourly steam gas (lbs)
    4:  'b1_econ_in',       # economizer inlet temp (F)
    5:  'b1_econ_out',      # economizer outlet temp (F)
    6:  'b1_runtime',       # runtime (hrs)
    # Boiler 2 (cols 7-11)
    7:  'b2_gas',
    8:  'b2_steam_gas',
    9:  'b2_econ_in',
    10: 'b2_econ_out',
    11: 'b2_runtime',
    # Boiler 3 (cols 12-21) — dual fuel
    12: 'b3_gas',
    13: 'b3_oil',           # oil (lbs)
    14: 'b3_steam_gas',
    15: 'b3_steam_oil',
    16: 'b3_econ_in',
    17: 'b3_econ_out',
    18: 'b3_steam_total',
    19: 'b3_runtime_gas',
    20: 'b3_runtime_oil',
    21: 'b3_runtime_total',
    # Boiler 4 (cols 22-31) — dual fuel
    22: 'b4_gas',
    23: 'b4_oil',
    24: 'b4_steam_gas',
    25: 'b4_steam_oil',
    26: 'b4_econ_in',
    27: 'b4_econ_out',
    28: 'b4_steam_total',
    29: 'b4_runtime_gas',
    30: 'b4_runtime_oil',
    31: 'b4_runtime_total',
    # Boiler 5 (cols 32-41) — dual fuel
    32: 'b5_gas',
    33: 'b5_oil',
    34: 'b5_steam_gas',
    35: 'b5_steam_oil',
    36: 'b5_econ_in',
    37: 'b5_econ_out',
    38: 'b5_steam_total',
    39: 'b5_runtime_gas',
    40: 'b5_runtime_oil',
    41: 'b5_runtime_total',
    # Plant aggregates (cols 42-48)
    42: 'casp_gas',
    43: 'gwsp_gas',
    44: 'campus_gas',
    45: 'gwsp_oil',
    46: 'casp_steam',
    47: 'gwsp_steam',
    48: 'campus_steam',
    # Tunnel steam (col 49)
    49: 'gwsp_tunnel_steam',    # kpph
    # Fuel oil tank (cols 50-52)
    50: 'fuel_oil_level',       # inH2O
    51: 'fuel_oil_gal',         # gallons total
    52: 'fuel_oil_change',      # change in tank (gal)
    # Water (cols 53-60)
    53: 'gwsp_raw_water',       # gal x1000
    54: 'gwsp_potable',         # gal x100
    55: 'gwsp_makeup_temp_c',   # Celsius
    56: 'gwsp_makeup_temp_f',   # Fahrenheit
    57: 'gwsp_permeate',
    58: 'casp_east_water',      # gal x100
    59: 'casp_west_water',      # gal x100
    60: 'gwsp_waste_water',
    # Generators (cols 61-70)
    61: 'gen1_runtime',
    62: 'gen1_gas',
    63: 'gen1_kwh',
    64: 'gen2_runtime',
    65: 'gen2_gas',
    66: 'gen2_kwh',
    67: 'gen3_fuel_rate',
    68: 'gen3_runtime',
    69: 'gen3_kwh',
    70: 'gen3_fuel_gal',
}

# Hourly sheet has 2 extra columns (B1 and B2 each get "hourly total steam")
# B1: cols 2-7 (extra col 6 = b1_steam_total), B2: cols 8-13 (extra col 12 = b2_steam_total)
# Everything after shifts by +2
HOURLY_COLS = {
    2:  'b1_gas',
    3:  'b1_steam_gas',
    4:  'b1_econ_in',
    5:  'b1_econ_out',
    6:  'b1_steam_total',   # HOURLY ONLY
    7:  'b1_runtime',
    8:  'b2_gas',
    9:  'b2_steam_gas',
    10: 'b2_econ_in',
    11: 'b2_econ_out',
    12: 'b2_steam_total',   # HOURLY ONLY
    13: 'b2_runtime',
    # Boiler 3 (cols 14-23)
    14: 'b3_gas',
    15: 'b3_oil',
    16: 'b3_steam_gas',
    17: 'b3_steam_oil',
    18: 'b3_econ_in',
    19: 'b3_econ_out',
    20: 'b3_steam_total',
    21: 'b3_runtime_gas',
    22: 'b3_runtime_oil',
    23: 'b3_runtime_total',
    # Boiler 4 (cols 24-33)
    24: 'b4_gas',
    25: 'b4_oil',
    26: 'b4_steam_gas',
    27: 'b4_steam_oil',
    28: 'b4_econ_in',
    29: 'b4_econ_out',
    30: 'b4_steam_total',
    31: 'b4_runtime_gas',
    32: 'b4_runtime_oil',
    33: 'b4_runtime_total',
    # Boiler 5 (cols 34-43)
    34: 'b5_gas',
    35: 'b5_oil',
    36: 'b5_steam_gas',
    37: 'b5_steam_oil',
    38: 'b5_econ_in',
    39: 'b5_econ_out',
    40: 'b5_steam_total',
    41: 'b5_runtime_gas',
    42: 'b5_runtime_oil',
    43: 'b5_runtime_total',
    # Plant aggregates (cols 44-50)
    44: 'casp_gas',
    45: 'gwsp_gas',
    46: 'campus_gas',
    47: 'gwsp_oil',
    48: 'casp_steam',
    49: 'gwsp_steam',
    50: 'campus_steam',
    # Tunnel steam (col 51)
    51: 'gwsp_tunnel_steam',
    # Fuel oil tank (cols 52-54)
    52: 'fuel_oil_level',
    53: 'fuel_oil_gal',
    54: 'fuel_oil_change',
    # Water (cols 55-62)
    55: 'gwsp_raw_water',
    56: 'gwsp_potable',
    57: 'gwsp_makeup_temp_c',
    58: 'gwsp_makeup_temp_f',
    59: 'gwsp_permeate',
    60: 'casp_east_water',
    61: 'casp_west_water',
    62: 'gwsp_waste_water',
    # Generators (cols 63-72)
    63: 'gen1_runtime',
    64: 'gen1_gas',
    65: 'gen1_kwh',
    66: 'gen2_runtime',
    67: 'gen2_gas',
    68: 'gen2_kwh',
    69: 'gen3_fuel_rate',
    70: 'gen3_runtime',
    71: 'gen3_kwh',
    72: 'gen3_fuel_gal',
}

# Keys that should be AVERAGED (temperatures), not summed
AVG_KEYS = {
    'b1_econ_in', 'b1_econ_out', 'b2_econ_in', 'b2_econ_out',
    'b3_econ_in', 'b3_econ_out', 'b4_econ_in', 'b4_econ_out',
    'b5_econ_in', 'b5_econ_out',
    'gwsp_makeup_temp_c', 'gwsp_makeup_temp_f',
    'gwsp_tunnel_steam',  # kpph is a rate, average it
    'gen3_fuel_rate',     # fuel rate is a rate
    'fuel_oil_level',     # tank level — average
}

# Keys that should take MAX (cumulative counters — runtime)
MAX_KEYS = {
    'b1_runtime', 'b2_runtime',
    'b3_runtime_gas', 'b3_runtime_oil', 'b3_runtime_total',
    'b4_runtime_gas', 'b4_runtime_oil', 'b4_runtime_total',
    'b5_runtime_gas', 'b5_runtime_oil', 'b5_runtime_total',
    'gen1_runtime', 'gen2_runtime', 'gen3_runtime',
}

# Column metadata for the JSON
def build_column_meta(col_map):
    equipment_map = {
        'b1': 'Boiler 1', 'b2': 'Boiler 2', 'b3': 'Boiler 3',
        'b4': 'Boiler 4', 'b5': 'Boiler 5',
        'casp': 'CASP', 'gwsp': 'GWSP', 'campus': 'Campus',
        'gen1': 'Generator 1', 'gen2': 'Generator 2', 'gen3': 'Generator 3',
        'fuel_oil': 'Fuel Oil Tank',
    }
    unit_hints = {
        'gas': 'scf', 'oil': 'lbs', 'steam': 'lbs', 'runtime': 'hrs',
        'econ_in': 'F', 'econ_out': 'F', 'temp_c': 'C', 'temp_f': 'F',
        'water': 'gal', 'potable': 'gal', 'permeate': 'gal', 'waste': 'gal',
        'kwh': 'kWh', 'fuel_gal': 'gal', 'fuel_rate': 'gal/hr',
        'level': 'inH2O', 'change': 'gal', 'tunnel': 'kpph',
    }
    columns = []
    for idx in sorted(col_map.keys()):
        key = col_map[idx]
        # Determine equipment prefix
        prefix = key.split('_')[0]
        if prefix + '_' + key.split('_')[1] in equipment_map:
            prefix = prefix + '_' + key.split('_')[1]
        equip = equipment_map.get(prefix, prefix.upper())
        # Determine unit
        unit = 'unknown'
        for hint, u in unit_hints.items():
            if hint in key:
                unit = u
                break
        columns.append({'key': key, 'equipment': equip, 'unit': unit})
    return columns


def safe_float(val):
    """Convert a cell value to float, returning 0 for None/invalid."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_date(val):
    """Parse a date cell value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y'):
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def parse_datetime(val):
    """Parse a datetime cell value to YYYY-MM-DDTHH:MM string."""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%dT%H:%M')
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M'):
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%dT%H:%M')
            except ValueError:
                continue
    return None


def read_daily_sheet(ws):
    """Read the Daily Data sheet and return list of row dicts."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Skip header rows (rows where col 0 is not numeric)
        day_val = row[0] if len(row) > 0 else None
        if day_val is None:
            continue
        try:
            int(float(day_val))
        except (ValueError, TypeError):
            continue

        date_str = parse_date(row[1]) if len(row) > 1 else None
        if not date_str:
            continue

        record = {'date': date_str}
        for col_idx, key in DAILY_COLS.items():
            if col_idx < len(row):
                record[key] = safe_float(row[col_idx])
            else:
                record[key] = 0.0
        rows.append(record)
    return rows


def read_hourly_sheet(ws):
    """Read the Hourly Data sheet and return list of row dicts."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        day_val = row[0] if len(row) > 0 else None
        if day_val is None:
            continue
        try:
            int(float(day_val))
        except (ValueError, TypeError):
            continue

        dt_str = parse_datetime(row[1]) if len(row) > 1 else None
        if not dt_str:
            continue

        record = {'datetime': dt_str}
        for col_idx, key in HOURLY_COLS.items():
            if col_idx < len(row):
                record[key] = safe_float(row[col_idx])
            else:
                record[key] = 0.0
        rows.append(record)
    return rows


def aggregate_monthly(daily_rows):
    """Aggregate daily rows into monthly rollups."""
    months = {}
    for row in daily_rows:
        month_key = row['date'][:7]  # YYYY-MM
        if month_key not in months:
            months[month_key] = []
        months[month_key].append(row)

    monthly = []
    all_keys = set(DAILY_COLS.values())

    for month_key in sorted(months.keys()):
        days = months[month_key]
        record = {'month': month_key}

        for key in all_keys:
            values = [d[key] for d in days if d.get(key) is not None]
            nonzero = [v for v in values if v != 0]

            if key in AVG_KEYS:
                # Average, excluding zeros (sensor off)
                record[key] = round(sum(nonzero) / len(nonzero), 2) if nonzero else 0
            elif key in MAX_KEYS:
                # Max (cumulative counter)
                record[key] = round(max(values), 2) if values else 0
            else:
                # Sum
                record[key] = round(sum(values), 2)

        monthly.append(record)

    return monthly


def main():
    input_file = 'WSU_Energy_Data_FY21-FY26_Combined.xlsx'
    if not os.path.exists(input_file):
        print(f"ERROR: {input_file} not found in current directory")
        sys.exit(1)

    print(f"Reading {input_file}...")
    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)

    # Daily data
    print("Processing Daily Data sheet...")
    ws_daily = wb['Daily Data']
    daily_rows = read_daily_sheet(ws_daily)
    print(f"  {len(daily_rows)} daily rows")

    # Monthly rollups
    print("Aggregating monthly rollups...")
    monthly_rows = aggregate_monthly(daily_rows)
    print(f"  {len(monthly_rows)} months")

    # Date range
    dates = [r['date'] for r in daily_rows]
    date_range = [min(dates), max(dates)] if dates else ['', '']

    # Source files list (just the one merged file)
    source_files = [input_file]

    # Build metadata
    meta = {
        'lastUpdated': datetime.now().strftime('%Y-%m-%d'),
        'dateRange': date_range,
        'monthCount': len(monthly_rows),
        'dayCount': len(daily_rows),
        'sourceFiles': source_files,
    }

    columns = build_column_meta(DAILY_COLS)

    # Write daily+monthly JSON
    output_daily = {
        'meta': meta,
        'columns': columns,
        'monthly': monthly_rows,
        'daily': daily_rows,
    }

    os.makedirs('data', exist_ok=True)
    daily_path = 'data/steam_plant.json'
    print(f"Writing {daily_path}...")
    with open(daily_path, 'w') as f:
        json.dump(output_daily, f, separators=(',', ':'))
    size_mb = os.path.getsize(daily_path) / (1024 * 1024)
    print(f"  {size_mb:.1f} MB")

    # Hourly data
    print("Processing Hourly Data sheet...")
    ws_hourly = wb['Hourly Data']
    hourly_rows = read_hourly_sheet(ws_hourly)
    print(f"  {len(hourly_rows)} hourly rows")

    hourly_columns = build_column_meta(HOURLY_COLS)
    meta_hourly = {
        'lastUpdated': meta['lastUpdated'],
        'dateRange': date_range,
        'monthCount': len(monthly_rows),
        'hourlyCount': len(hourly_rows),
        'sourceFiles': source_files,
    }

    output_hourly = {
        'meta': meta_hourly,
        'columns': hourly_columns,
        'hourly': hourly_rows,
    }

    hourly_path = 'data/steam_plant_hourly.json'
    print(f"Writing {hourly_path}...")
    with open(hourly_path, 'w') as f:
        json.dump(output_hourly, f, separators=(',', ':'))
    size_mb = os.path.getsize(hourly_path) / (1024 * 1024)
    print(f"  {size_mb:.1f} MB")

    wb.close()

    print(f"\nDone! Summary:")
    print(f"  Date range: {date_range[0]} to {date_range[1]}")
    print(f"  {len(daily_rows)} daily rows, {len(monthly_rows)} months")
    print(f"  {len(hourly_rows)} hourly rows")
    print(f"  Files: {daily_path}, {hourly_path}")


if __name__ == '__main__':
    main()
